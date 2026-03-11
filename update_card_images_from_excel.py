# update_card_images.py
import os
import django
import openpyxl
import glob

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from pricehub.models import Card, OnePieceCard


def update_pokemon_card_images():
    """포켓몬 카드 이미지 URL 업데이트"""
    print("\n" + "=" * 80)
    print("🎴 포켓몬 카드 이미지 URL 업데이트")
    print("=" * 80 + "\n")
    
    # 엑셀 파일 찾기
    excel_files = glob.glob("*.xlsx")
    
    if not excel_files:
        print("❌ 엑셀 파일을 찾을 수 없습니다.")
        return
    
    print(f"📂 발견된 엑셀 파일: {len(excel_files)}개\n")
    
    total_updated = 0
    total_not_found = 0
    total_already_has_url = 0
    
    for excel_file in excel_files:
        print(f"📄 처리 중: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            file_updated = 0
            file_not_found = 0
            file_already_has_url = 0
            
            # 6행부터 데이터 시작 (1~5행은 헤더)
            for row_idx, row in enumerate(ws.iter_rows(min_row=6), start=6):
                try:
                    # B열: 상품코드 (인덱스 1)
                    shop_product_code = row[1].value
                    
                    # U열: 이미지 URL (인덱스 25)
                    image_url = row[25].value if len(row) > 25 else None
                    
                    if not shop_product_code:
                        continue
                    
                    # 상품코드 문자열로 변환
                    shop_product_code = str(shop_product_code).strip()
                    
                    # 이미지 URL이 없으면 스킵
                    if not image_url:
                        continue
                    
                    image_url = str(image_url).strip()
                    
                    # DB에서 카드 찾기
                    try:
                        card = Card.objects.get(shop_product_code=shop_product_code)
                        
                        # 이미 이미지 URL이 있으면 스킵 (선택사항)
                        if card.image_url:
                            file_already_has_url += 1
                            print(f"  [행 {row_idx}] ⏭️  이미 이미지 존재: {shop_product_code}")
                            continue
                        
                        # 이미지 URL 업데이트
                        card.image_url = image_url
                        card.save()
                        
                        file_updated += 1
                        print(f"  [행 {row_idx}] ✅ 업데이트: {shop_product_code}")
                        
                    except Card.DoesNotExist:
                        file_not_found += 1
                        print(f"  [행 {row_idx}] ❌ 카드 없음: {shop_product_code}")
                
                except Exception as e:
                    print(f"  [행 {row_idx}] ⚠️  오류: {e}")
                    continue
            
            print(f"  ✅ 업데이트: {file_updated}개")
            print(f"  ⏭️  이미 존재: {file_already_has_url}개")
            print(f"  ❌ 찾을 수 없음: {file_not_found}개")
            print()
            
            total_updated += file_updated
            total_not_found += file_not_found
            total_already_has_url += file_already_has_url
            
        except Exception as e:
            print(f"  ❌ 파일 처리 오류: {e}\n")
            continue
    
    # 전체 결과
    print("=" * 80)
    print("📊 전체 결과")
    print("=" * 80)
    print(f"✅ 총 업데이트: {total_updated}개")
    print(f"⏭️  이미 존재: {total_already_has_url}개")
    print(f"❌ 찾을 수 없음: {total_not_found}개")
    print()


def update_onepiece_card_images():
    """원피스 카드 이미지 URL 업데이트"""
    print("\n" + "=" * 80)
    print("🏴‍☠️ 원피스 카드 이미지 URL 업데이트")
    print("=" * 80 + "\n")
    
    # 엑셀 파일 찾기
    excel_files = glob.glob("*.xlsx")
    
    if not excel_files:
        print("❌ 엑셀 파일을 찾을 수 없습니다.")
        return
    
    print(f"📂 발견된 엑셀 파일: {len(excel_files)}개\n")
    
    total_updated = 0
    total_not_found = 0
    total_already_has_url = 0
    
    for excel_file in excel_files:
        print(f"📄 처리 중: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            file_updated = 0
            file_not_found = 0
            file_already_has_url = 0
            
            # 6행부터 데이터 시작
            for row_idx, row in enumerate(ws.iter_rows(min_row=6), start=6):
                try:
                    # B열: 상품코드
                    shop_product_code = row[1].value
                    
                    # U열: 이미지 URL
                    image_url = row[25].value if len(row) > 25 else None
                    
                    if not shop_product_code:
                        continue
                    
                    shop_product_code = str(shop_product_code).strip()
                    
                    if not image_url:
                        continue
                    
                    image_url = str(image_url).strip()
                    
                    # DB에서 카드 찾기
                    try:
                        card = OnePieceCard.objects.get(shop_product_code=shop_product_code)
                        
                        # 이미 이미지 URL이 있으면 스킵
                        if card.image_url:
                            file_already_has_url += 1
                            print(f"  [행 {row_idx}] ⏭️  이미 이미지 존재: {shop_product_code}")
                            continue
                        
                        # 이미지 URL 업데이트
                        card.image_url = image_url
                        card.save()
                        
                        file_updated += 1
                        print(f"  [행 {row_idx}] ✅ 업데이트: {shop_product_code}")
                        
                    except OnePieceCard.DoesNotExist:
                        file_not_found += 1
                        print(f"  [행 {row_idx}] ❌ 카드 없음: {shop_product_code}")
                
                except Exception as e:
                    print(f"  [행 {row_idx}] ⚠️  오류: {e}")
                    continue
            
            print(f"  ✅ 업데이트: {file_updated}개")
            print(f"  ⏭️  이미 존재: {file_already_has_url}개")
            print(f"  ❌ 찾을 수 없음: {file_not_found}개")
            print()
            
            total_updated += file_updated
            total_not_found += file_not_found
            total_already_has_url += file_already_has_url
            
        except Exception as e:
            print(f"  ❌ 파일 처리 오류: {e}\n")
            continue
    
    # 전체 결과
    print("=" * 80)
    print("📊 전체 결과")
    print("=" * 80)
    print(f"✅ 총 업데이트: {total_updated}개")
    print(f"⏭️  이미 존재: {total_already_has_url}개")
    print(f"❌ 찾을 수 없음: {total_not_found}개")
    print()


def update_all_card_images():
    """포켓몬 + 원피스 카드 이미지 URL 모두 업데이트"""
    print("\n" + "=" * 80)
    print("🎴 모든 카드 이미지 URL 업데이트")
    print("=" * 80)
    
    update_pokemon_card_images()
    update_onepiece_card_images()


if __name__ == '__main__':
    print("\n🖼️  카드 이미지 URL 업데이트 도구")
    print("=" * 80)
    print("\n선택하세요:")
    print("  1. 포켓몬 카드 이미지 업데이트")
    print("  2. 원피스 카드 이미지 업데이트")
    print("  3. 모든 카드 이미지 업데이트")
    print("  4. 종료")
    
    choice = input("\n선택 (1/2/3/4): ").strip()
    
    if choice == '1':
        confirm = input("포켓몬 카드 이미지를 업데이트하시겠습니까? (yes/no): ")
        if confirm.lower() == 'yes':
            update_pokemon_card_images()
    
    elif choice == '2':
        confirm = input("원피스 카드 이미지를 업데이트하시겠습니까? (yes/no): ")
        if confirm.lower() == 'yes':
            update_onepiece_card_images()
    
    elif choice == '3':
        confirm = input("모든 카드 이미지를 업데이트하시겠습니까? (yes/no): ")
        if confirm.lower() == 'yes':
            update_all_card_images()
    
    elif choice == '4':
        print("종료합니다.")
    
    else:
        print("❌ 잘못된 선택입니다.")
